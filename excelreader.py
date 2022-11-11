from openpyxl import load_workbook
import os


def xlsxHandler(index):
    wb = load_workbook(filename = os.getcwd() + r'\places.xlsx', read_only=True)
    # sheets_name = wb.sheetnames
    sheets = wb.worksheets
    print(len(sheets))
    for i in range(len(sheets)):
        print(i+1,end=" ")
        print(sheets[i].title)
    # sheetnames = [sheet.title for sheet in sheets]
    # for sheet in sheets:
    data_from_xl = sheets[index]._cells_by_row(min_col=2,max_col=3,min_row=2,max_row=200)
    print(data_from_xl)
    places_kw = [sheets[index].title]
    for data in data_from_xl:
        if data[0].value and data[1].value:
            places_kw.append(data[1].value + " " + data[0].value)
    print(len(places_kw))
    print(places_kw)

if __name__ == "__main__":
    xlsxHandler(0)