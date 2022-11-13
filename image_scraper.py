from multiprocessing import freeze_support
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait  #..............
from selenium.webdriver.support import expected_conditions as EC    #................
import requests # to get image from the web
import shutil # to save it locally
from urllib.parse import urlparse
# from bs4 import BeautifulSoup
# import xlsxwriter
import time
import func_timeout
import multiprocessing
# from selenium.webdriver.common.proxy import Proxy, ProxyType  #.................
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.chrome.service import Service
# from webdriver_manager.chrome import ChromeDriverManager
# from msvcrt import getche, getch
from openpyxl import load_workbook
import os
import sys

# url = "https://www.google.com/search?as_st=y&tbm=isch&as_q=&as_epq=badshahi+mosque&as_oq=&as_eq=&cr=&as_sitesearch=&safe=images&tbs=isz:lt,islt:xga"


def logger(state):
    if state == "on":
        log_dir = os.getcwd() + r"\logs"
        log_file = log_dir + r"\image_scraper.log"
        # Cheks if folder is there or not
        if not os.path.exists(log_dir):
            # shutil.rmtree(log_dir)
            os.makedirs(log_dir)
        # Cheks if file is already there or not
        if os.path.exists(log_file):
            os.remove(log_file) # one file at a time
        # Open file to start writing
        sys.stdout = open(log_file, "w")
    elif state == "off":
        print("----------------------End of File----------------------")
        sys.stdout.close()

class XlsxManager:
    def __init__(self):
        self.wb = load_workbook(filename = os.getcwd() + r'\places.xlsx', read_only=True)
        self.sheets = self.wb.worksheets
        # self.search_kw_list = None

    def dirManager(self, dir, mode=None):
        if mode == "replace":
            if os.path.exists(dir):
                shutil.rmtree(dir)
            os.makedirs(dir)
        else:
            if not os.path.exists(dir):
                os.makedirs(dir)

    def xlsxManager(self):
        # imgScraper1 = ImagesScraper()
        # imgScraper2 = ImagesScraper()
        self.dirManager('Images','replace')
        if len(self.sheets)%2 == 0:
            for index in range(0, len(self.sheets), 2):
                search_kw_list1 = self.sheetReader(index)
                search_kw_list2 = self.sheetReader(index+1)
                p1 = multiprocessing.Process(target=sheetManager, args=(index, len(self.sheets), search_kw_list1,))
                p2 = multiprocessing.Process(target=sheetManager, args=(index+1, len(self.sheets), search_kw_list2,))
                p1.start()
                print("P1 started")
                p2.start()
                print("P2 started")
                p1.join()
                p2.join()
        else:
            for index in range(0, len(self.sheets), 2):
                # sheets = self.sheets
                if len(self.sheets)-1 == index:
                    search_kw_list1 = self.sheetReader(index)
                    p1 = multiprocessing.Process(target=sheetManager, args=(index, len(self.sheets), search_kw_list1,))
                    p1.start()
                    print("P1 started")
                    p1.join()
                else:
                    search_kw_list1 = self.sheetReader(index)
                    search_kw_list2 = self.sheetReader(index+1)
                    p1 = multiprocessing.Process(target=sheetManager, args=(index, len(self.sheets), search_kw_list1,))
                    p2 = multiprocessing.Process(target=sheetManager, args=(index+1, len(self.sheets), search_kw_list2,))
                    p1.start()
                    print("P1 started")
                    p2.start()
                    print("P2 started")
                    p1.join()
                    p2.join()

    def sheetReader(self, index):
        # sheets_name = wb.sheetnames
        # sheets = self.wb.worksheets
        # sheetnames = [sheet.title for sheet in sheets]
        # for sheet in sheets:
        data_from_sheet = self.sheets[index]._cells_by_row(min_col=2,max_col=3,min_row=2,max_row=200)
        print(data_from_sheet)
        places_kw_list = [self.sheets[index].title]
        for data in data_from_sheet:
            if data[0].value and data[1].value:
                places_kw_list.append((data[0].value, data[1].value))
        
        return places_kw_list  # Return
            
class sheetManager:
    def __init__(self, index, sheets_len, search_kw_list):
        print("\n\n----------------------------------------------------------------------------------------")
        print("[{}/{}] Scraping List of Places from: {}".format(index+1,sheets_len,search_kw_list[0]))
        # imgScraper1.scraper_manager(search_kw_list)
        process_div = len(search_kw_list)//6
        search_kw_list1 = [x for x in search_kw_list[:process_div]]
        search_kw_list2 = [search_kw_list[0]] + [x for x in search_kw_list[process_div:process_div*2]]
        search_kw_list3 = [search_kw_list[0]] + [x for x in search_kw_list[process_div*2:process_div*3]]
        search_kw_list4 = [search_kw_list[0]] + [x for x in search_kw_list[process_div*3:process_div*4]]
        search_kw_list5 = [search_kw_list[0]] + [x for x in search_kw_list[process_div*4:process_div*5]]
        search_kw_list6 = [search_kw_list[0]] + [x for x in search_kw_list[process_div*5:]]
        sp1 = multiprocessing.Process(target=ImagesScraper, args=(search_kw_list1,))
        sp2 = multiprocessing.Process(target=ImagesScraper, args=(search_kw_list2,))
        sp3 = multiprocessing.Process(target=ImagesScraper, args=(search_kw_list3,))
        sp4 = multiprocessing.Process(target=ImagesScraper, args=(search_kw_list4,))
        sp5 = multiprocessing.Process(target=ImagesScraper, args=(search_kw_list5,))
        sp6 = multiprocessing.Process(target=ImagesScraper, args=(search_kw_list6,))
        sp1.start()
        print("SP1 started")
        sp2.start()
        print("SP2 started")
        sp3.start()
        print("SP3 started")
        sp4.start()
        print("SP4 started")
        sp5.start()
        print("SP5 started")
        sp6.start()
        print("SP6 started")
        sp1.join()
        sp2.join()
        sp3.join()
        sp4.join()
        sp5.join()
        sp6.join()
        # for kw in search_kw_list[1:]:  # [1:] Coz first element is name of country
        #     print("-----------------")
        #     search_querry = self.search_querry_manager(kw)
        #     print("Search Keyword: {}".format(search_querry))
        #     self.scraper_manager(kw)  # Parsing kw instead of search_querry coz life happens............


class ImagesScraper:
    def __init__(self, search_kw_list):
        # logger("on") # Turning on the logger ;)
        self.driver = self.gen_driver()
        self.dirManager('Images')
        self.search_kw_list = search_kw_list
        self.search_querry = None
        self.imgurls = None
        self.image_quantity = 4

        self.scraper_manager()

    def gen_driver(self):
        chrome_options = uc.ChromeOptions()
        chrome_options.headless = True
        # chrome_options.add_argument('--proxy-server=http://'+PROXY)
        driver = uc.Chrome(options=chrome_options)
        return driver  

    def search_querry_manager(self, kw):
        if kw[0].lower() in kw[1].lower():
            return kw[1]
        else:
            return kw[1] + " " + kw[0]

    def scraper_manager(self):
        # self.search_kw_list = search_kw_list
        for i, kw in enumerate(self.search_kw_list[1:],1):   # [1:] Coz first element is name of country
            search_querry = self.search_querry_manager(kw)
            print("-----------------")
            print("[{}/{}] Search Keyword: {}".format(i,len(self.search_kw_list)-1,search_querry))

            self.search(search_querry)
            self.fetch_image_urls(self.image_quantity)
            try:
                self.dowload_images(kw)
                # func_timeout.func_timeout(30, self.dowload_images, args=[kw])
            # except func_timeout.FunctionTimedOut:
            #     print('-./>>>>>>>>>/-/.-:: 30s completed but process did not complete. Skipping>>>>')
            except Exception as e:
                print("Coudn't dowload image, something went wrong: {}\nSkipping>>>>>>>".format(e))

    def search(self, search_querry):
        self.search_querry = search_querry
        querry = self.search_querry.replace(" ", "+")

        # HD Quality
        # self.driver.get("https://www.google.com/search?as_st=y&tbm=isch&as_q=&as_epq={}&as_oq=&as_eq=&cr=&as_sitesearch=&safe=images&tbs=isz:lt,islt:xga".format(querry))
        
        # HD + Wide
        # self.driver.get("https://www.google.com/search?as_st=y&tbm=isch&as_q=&as_epq={}&as_oq=&as_eq=&cr=&as_sitesearch=&safe=images&tbs=isz:lt,islt:xga,iar:w".format(querry))
        
        # HD + Type:Photo
        self.driver.get("https://www.google.com/search?as_st=y&tbm=isch&as_q=&as_epq={}&as_oq=&as_eq=&cr=&as_sitesearch=&safe=images&tbs=isz:lt,islt:xga,itp:photo".format(querry))
        
        # HD + Wide + Type:Photo
        # self.driver.get("https://www.google.com/search?as_st=y&tbm=isch&as_q=&as_epq={}&as_oq=&as_eq=&cr=&as_sitesearch=&safe=images&tbs=isz:lt,islt:xga,itp:photo,iar:w".format(querry))


    def fetch_image_urls(self, quantity):
        # imgurle = self.driver.find_element_by_xpath('//*[@id="islrg"]/div[1]/div[%s]/a[1]/div[1]/img'%(str(indx)))
        # driver.find_elements(By.CSS_SELECTOR, '#__ZONE__main > div > div > div.main-categories.lohp-row.max-width-container > ul > li > a')
        imgurls = []
        for i in range(1,quantity+1):
            try:
                img_card = self.driver.find_element(By.XPATH, '//*[@id="islrg"]/div[1]/div[%s]/a[1]/div[1]/img'%(str(i)))
                img_card.click()
                time.sleep(6)
                popup_class = self.driver.find_elements(By.CLASS_NAME, 'n3VNCb')
                for popup in popup_class:
                    imgurl = popup.get_attribute("src")
                    if "http" in  imgurl and not "encrypted" in imgurl and not "base64" in imgurl:
                        imgurls.append(imgurl)
                        print("[{}] Fetched Image url: {}".format(i,imgurl))
            except Exception as e:
                print("Error During URL Fetch: {}\nSkipping>>>>>>>".format(e))
                continue

        self.imgurls = imgurls    # Return

    def dirManager(self, dir, mode=None):
        if mode == "replace":
            if os.path.exists(dir):
                shutil.rmtree(dir)
            os.makedirs(dir)
        else:
            if not os.path.exists(dir):
                os.makedirs(dir)

    def dowload_images(self, kw):
        # Creates a new folder to save the images. If already exists, Overwrites it.
        images_dir = 'Images'
        sub_img_dir = 'Images\\' + self.search_kw_list[0]
        sub_sub_img_dir = sub_img_dir + '\\' + kw[0]
        sub_sub_sub_img_dir = sub_sub_img_dir + '\\' + self.search_querry
        self.dirManager(sub_img_dir)
        self.dirManager(sub_sub_img_dir)
        self.dirManager(sub_sub_sub_img_dir)
        print("Saving Images at: {}".format(sub_sub_img_dir))

        downloaded_images = []

        def requestsGet(image_url):
            return requests.get(image_url, stream = True)

        for i, image_url in enumerate(self.imgurls,1):
            try:
                #extact filename without extension from URL
                extract_url = urlparse(image_url)
            
                ## Set up the image URL and filename
                # filename = os.path.basename(extract_url.path)
                filename, file_extension = os.path.splitext(os.path.basename(extract_url.path))
                if file_extension:
                    filename = self.search_querry + " " + str(i) + file_extension
                else:
                    filename = self.search_querry + " " + str(i) + ".jpg"

                # if similar name exists, don't overwrite please
                # if filename in downloaded_images:
                #     filename = str(i) + filename
                # else:
                #     downloaded_images.append(filename)

                try:
                    # Open the url image, set stream to True, this will return the stream content.
                    r = func_timeout.func_timeout(30, requestsGet, args=[image_url])
                except func_timeout.FunctionTimedOut:
                    print('-./>>>>>>>>>/-/.-:: 30s completed but no requests were gotten. Skipping>>>>')
                    raise Exception("30s Timeout!")
                # r = requestsGet(image_url)

                # Check if the image was retrieved successfully
                if r.status_code == 200:
                    # Set decode_content value to True, otherwise the downloaded image file's size will be zero.
                    r.raw.decode_content = True
                    
                    # Open a local file with wb ( write binary ) permission.
                    save_image_dir = os.path.join(os.getcwd(), os.path.join(sub_sub_sub_img_dir, filename))
                    with open(save_image_dir,'wb') as f:
                        try:
                            # shutil.copyfileobj(r.raw, f)
                            func_timeout.func_timeout(30, shutil.copyfileobj, args=[r.raw, f])
                        except func_timeout.FunctionTimedOut:
                            print("-././-./-.-/-/.-:: 30s completed but image could'nt be downloaded. Skipping>>>>")
                            raise Exception("30s Timeout!")
                        
                    print('[{}] Image sucessfully Downloaded: {}'.format(i, filename))
                else:
                    print('[{}] [Status Code: {}] Image Couldn\'t be retreived: {}'.format(i, r.status_code, image_url))
            except Exception as e:
                    print("Error During Downloading: {}\nSkipping>>>>>>>".format(e))
                    continue

        # logger("off") # Turning logger off ;(

 
if __name__ == "__main__":
    start_time = time.time()
    scrapeman = XlsxManager()
    scrapeman.xlsxManager()
    end_time = time.time() - start_time
    print("--- %s seconds ---" % (end_time))
    print("Program took: {} minutes".format(end_time/60))
    # print("Enter the search term: ", end="")
    # # scrapeman.search_querry = input()
    # scrapeman.search(input())
    # scrapeman.fetch_image_urls(8)
    # # print(imgurls)
    # scrapeman.dowload_images()
    # # search_querry.replace(" ", "+")
    # # driver = gen_driver()
    # # driver.get("https://www.google.com/search?as_st=y&tbm=isch&as_q=&as_epq={}&as_oq=&as_eq=&cr=&as_sitesearch=&safe=images&tbs=isz:lt,islt:xga".format(search_querry))
    